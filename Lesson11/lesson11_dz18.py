import csv
import openpyxl

wb = openpyxl.Workbook()
wb.create_sheet(title="First list", index=0)
wb.remove(wb['Sheet'])
sheet = wb["First list"]
list_value = []
with open("dz_17.csv", encoding='utf-8') as file:
    file_reader = csv.reader(file, delimiter=",")
    for row in list(file_reader):
        list_value.append([row[0], row[1], row[3]])

for col_index, column in enumerate(list_value):
    if col_index == 0:
        continue
    else:
        cell = sheet.cell(row=1, column=col_index + 1)
        cell.value = f"Person {col_index }"

for col_index, column in enumerate(list_value):
    for row_index, value in enumerate(column):
        cell = sheet.cell(row=row_index+2, column=col_index+1)
        cell.value = value

wb.save('dz_18.xlsx')


