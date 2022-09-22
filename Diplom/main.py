import openpyxl
from class_person import Person


def write_id_person():
    name_column = ["Ім'я", "Прізвище", "По батькові", "День народження", "День смерті", "Стать", "Вік"]
    name = input("Введіть ім'я: ").title()
    surname = input("Введіть прізвище: ").title()
    middle_name = input("Введіть по-батькові: ").title()
    day_of_birthday = input("Введіть дату народження: ")
    day_of_death = input("Введіть дату смерті (Щоб пропустити цей пункт натисніть Enter): ")
    sex = input("Введіть стать: ")

    person = Person(name, day_of_birthday, sex, day_of_death, surname, middle_name)

    try:
        wb = openpyxl.load_workbook('diplom.xlsx')
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        wb.create_sheet(title="First list", index=0)
        wb.remove(wb['Sheet'])
    finally:
        sheet = wb["First list"]
        rows = sheet.max_row
        for col_index, value in enumerate(name_column):
            cell = sheet.cell(row=1, column=col_index + 1)
            cell.value = value
        for col_index, value in enumerate(person.create_list()):
            cell = sheet.cell(row=rows + 1, column=col_index + 1)
            cell.value = value
    wb.save('diplom.xlsx')


def search_id_person():
    input_item_search = input("Введіть кого ви хочете знайти: ").replace(" ", "").lower()

    fields_rows = []
    wb = openpyxl.load_workbook('diplom.xlsx')
    sheet = wb["First list"]
    rows = sheet.max_row
    cols = sheet.max_column
    for i in range(1, rows+1):
        value_row = []
        for j in range(1, cols+1):
            cell = sheet.cell(row=i, column=j)
            value = str(cell.value)
            if i == 1:
                continue
            else:
                value_row.append(value)
        fields_rows.append(value_row)
    wb.save('diplom.xlsx')
    count = 0
    for rows in fields_rows:
        try:
            if "".join(rows[0:3]).lower().index(input_item_search) == 0 or \
                    "".join(rows[0:3]).lower().index(input_item_search) == 1:
                print(Person.print_list(rows))
                count += 1
                continue
        except ValueError:
            continue
    if count == 0:
        print("За вашим запитом нічого не знайдено.")


while True:
    function = {"1": write_id_person, "2": search_id_person}
    print("""Дана програма дозволяє працювати з даними про людей
* Якщо ви хочете додати дані про людину натисніть 1
* Якщо ви хочете знайти дані про якусь людину натисніть 2
* Якщо ви хочете вийти натисніть 3""")
    selection = input("Введіть ваш вибір: ")
    if not selection.isdigit() or int(selection) == 0 or int(selection) > 3:
        print("Ви ввели некоректне значення, спробуйте ще")
        print("-" * 50)
        continue
    elif int(selection) == 3:
        break
    function[selection]()
    print("-" * 50 + "\n")
